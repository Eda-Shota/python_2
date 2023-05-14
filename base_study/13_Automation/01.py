# Excelの仕事を自動化する
# openpyxlライブラリ
# Ecxelファイルに対応するライブラリの一つ

pip install openpyxl

import openpyxl
book = openpyxl.Workbook() # ワークブックを作成
book.save("catalog.xlsx") # ワークブックを保存

# セルの値を読み書きする
import openpyxl
book = openpyxl.Workbook()
sheet = book.active # ワークシートを取得
sheet["A1"] = "hat" # セルA1に値を書き込む
sheet["B1"] = 2000 # セルB1に値を書き込む
book.save("catalog.xlsx")

# 既存のExcelファイルを開く
import openpyxl
book = openpyxl.load_workbook("catalog.xlsx") # ワークブックを開く
sheet = book.active
print(sheet["A1"].value, sheet["B1".value])

# 複数のセルを読み書きする
import openpyxl
book = openpyxl.load_workbook("catalog.xlsx")
sheet = book.active
catalog = [("hat", 2000), ("shirt", 1000), ("socks", 500)] # 商品のリスト
for i, (name, price) in enumerate(catalog, 1): # 全商品を順に処理する
    sheet[f"A{i}"] = name # 名前を書き込む
    sheet[f"B{i}"] = price # 価格を書き込む
book.save("catalog.xlsx")

import openpyxl
book = openpyxl.load_workbook("catalog.xlsx")
sheet = book.active
for name, price in sheet.iter_rows(values_only=True): # 1行ずつ処理する
    print(name, price)

# Excelの作業をプログラムに肩代わりさせる
import openpyxl
book = openpyxl.load_workbook("catalog.xlsx")
sheet = book.active
total = 0 # 合計を初期化
for i, (name, price) in enumerate(sheet.iter_rows(values_only=True), 1): # 1行ずつ処理
    if not price: # 価格が空なら次の行へ
        continue
    if name == "Total": # 合計の行ならば終了
        break
    total += price # 会計に価格を加算
else:
    i += 2 # 会計を書き込む行番号
sheet[f"A{i}"] = "Total" # 列AにTotalと書き込む
sheet[f"B{i}"] = total # 列Bに合計を書き込む
book.save("catalog.xlsx")